import openpyxl


class ReadExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_excel(self):
        wb = openpyxl.load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(2, sheet.max_row + 1):
            sub_test = sheet.cell(i, 1).value
            test_data.append(sub_test)

        return test_data

if __name__ == '__main__':
    data = ReadExcel('T1data.xlsx', 'python').read_excel()
    for i in data:
        print(i)
