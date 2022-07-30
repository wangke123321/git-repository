import xlrd
import openpyxl
from common.log import Log
import os


class DoExcel:

    def __init__(self):
        self.log = Log()

    def get_excellist(self):
        file_list = []
        root_dir = os.getcwd()
        dir_or_files = os.listdir(root_dir)
        for dir_file in dir_or_files:
            dir_file_path = os.path.join(root_dir, dir_file)
            if os.path.isdir(dir_file_path):
                pass
            else:
                filepath, filename = os.path.split(dir_file_path)
                file_list.append(filename)
                case_list = list(filter(lambda case: ".xlsx" in case, file_list))
        if not case_list:
            raise Exception(self.log.logMsg(3, ("case not exist")))
        return case_list

    def openworkbook(self, xlsx):
        try:
            workbook = xlrd.open_workbook(xlsx)
            self.sheets = workbook.sheet_names()
            return workbook
        except Exception as e:
            self.log.logMsg(3, f'case not exist')
        return None

    def getname(self, name):
        sheetname = self.workbook.sheet_by_name(name)
        return sheetname

    def getsheetnames(self):
        sheetname = []
        for sheet in self.sheets:
            if sheet == "ElementMap":
                pass
            else:
                sheetname.append(sheet)
        return sheetname

    def get_value(self, sheet, row, col):
        table = self.workbook.sheet_by_name(sheet)
        cell_value = table.cell_value(row, col)
        return cell_value

    def enablesheet_list(self, xlsx):
        name_status = {}
        workbook = openpyxl.load_workbook(xlsx)
        sheet = workbook["CaseRunConfig"]
        sheetnames = self.getsheetnames()
        nrow = 0
        for sheetname in sheetnames:
            nrow += 1
            if sheetname == "CaseRunConfig":
                pass
            else:
                sheet.cell(nrow + 1, 1).value = sheetname
                name_status[sheetname] = sheet.cell(nrow + 1, 2).value
        return name_status

    def get_excel_title(self, xlsx):
        rows_obj = self.openworkbook(xlsx).sheet_by_name(self.sheet_name_title).get_rows()
        rows_objs = tuple(rows_obj)[0]
        title = []
        for r in rows_objs:
            title.append(r.value)
        return title

    def get_all_value(self, xlsx):
        self.workbook = xlrd.open_workbook(xlsx)
        self.sheets = self.workbook.sheet_names()
        values = []
        status_lists = self.enablesheet_list(xlsx)
        for st in self.sheets:
            if st in ("ElementMap", "CaseRunConfig"):
                pass
            else:
                if status_lists[st] == 1:
                    self.sheet_name_title = st
                    rows_obj = self.workbook.sheet_by_name(st).get_rows()
                    i = 0
                    for row_tuple in rows_obj:
                        self.value_list = []
                        n = 0
                        for val in row_tuple:
                            if val.value == None:
                                val.value = ""
                                n = n + 1
                            self.value_list.append(val.value)
                        if n >= 10:
                            pass
                        else:
                            if i == 0:
                                pass
                            else:
                                values.append(self.value_list)
                        i += 1
                else:
                    pass
        return values

    def get_listdict_all_value(self, xlsx):
        self.xlsx = xlsx
        all_values = self.get_all_value(xlsx)
        sheet_title = self.get_excel_title(xlsx)
        value_list = []
        for value in all_values:
            value_list.append(dict(zip(sheet_title, value)))
        return value_list

    def get_all_value_map(self):
        rows_obj = self.getname("ElementMap").get_rows()
        values = []
        for row_tuple in rows_obj:
            value_list = []
            n = 0
            for val in row_tuple:
                if val.value == None:
                    val.value = ""
                    n = n + 1
                value_list.append(val.value)
            if n >= 5:
                pass

            else:
                values.append(value_list)
        del (values[0])
        return values

    def get_excel_title_map(self):
        rows_obj = self.getname("ElementMap").get_rows()
        rows_objs = tuple(rows_obj)[0]
        title = []
        for r in rows_objs:
            title.append(r.value)
        return title

    def get_listdict_all_value_map(self):
        sheet_title = self.get_excel_title_map()
        all_values = self.get_all_value_map()
        value_list = []
        for value in all_values:
            value_list.append(dict(zip(sheet_title, value)))
        return value_list